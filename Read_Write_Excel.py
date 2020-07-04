#-*- coding: UTF-8 -*-
# 实现读写测试数据功能  -- lily 2020-6-30

from openpyxl import load_workbook
from openpyxl import Workbook

def read_casedata(filename,sheetname): #用于读取测试用例
    wb=load_workbook(filename)  #打开指定的工作薄
    sheet_1=wb[sheetname]       #定位表单
    all_case=[]
    for i in range(2,1+sheet_1.max_row): #从第2行开始读取
        data=[]
        for j in range(1,sheet_1.max_column-1): #
            data.append(sheet_1.cell(row=i,column=j).value)  #定位单元格，读取数据
        all_case.append(data)
    return all_case

def write_test_result(filename,sheetname,results,col=0):  #用于写测试结果
    #col指定列数：0 写入最后一列，1写入倒数第二列
    wb=load_workbook(filename)  #打开指定的工作薄
    sheet_1 = wb[sheetname]
    for i in range(2,1+sheet_1.max_row): #从第2行开始
        data=str(results[i-2])  #列表/字典类型不能直接写入excel, 因此要将其转换为字符串格式
        sheet_1.cell(row=i,column=(sheet_1.max_column-col)).value=data #定位单元格，读取数据
    wb.save(filename)

if __name__ == '__main__':
    all_case=read_casedata("QCD_API_casedata.xlsx","update_name")
    print(all_case)
    results=[{'pass':1},'pass','pass','pass','pass','no']
    write_test_result("QCD_API_casedata.xlsx","update_name", results)  #注意，写入前要确认excel是关闭状态
